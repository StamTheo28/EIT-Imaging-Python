import tkinter as tk
import webbrowser
from tkinter import StringVar, filedialog, messagebox, ttk

import matplotlib
import openpyxl
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from filehelpers import open_file_at_frequency

from visualisation import greit_visualisation


BUTTON1_DEFAULT_TEXT = "1. No Item Selected"
BUTTON2_DEFAULT_TEXT = "2. No Item Selected"
BASELINE_DEFAULT_TEXT = "None Set"


class VascuSensGUI:

    def __init__(self):
        # Initialise a TK Window
        self.window = tk.Tk()

        # Set configurations for the window object
        self.configure_window()

        # Place widgets and define layout for window object
        self.setup_layout()

    def configure_window(self):
        """ Configuration of the TK window object """

        # Configure matplotlib backend
        matplotlib.use("Agg")

        # Set window title and make non-resizable by the user
        self.window.title("VascuSens Blockage Visualiser - CS33")
        self.window.resizable(False, False)

        # Menubar Configuration
        menubar = tk.Menu(self.window)
        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Upload Item", command=self.upload_action)
        filemenu.add_command(label="Edit Baseline File Path", command=self.edit_baseline_path)
        filemenu.add_separator()
        filemenu.add_command(label="Exit", command=self.window.quit)
        menubar.add_cascade(label="File", menu=filemenu)
        helpmenu = tk.Menu(menubar, tearoff=0)
        helpmenu.add_command(
            label="User Guide",
            command=lambda: webbrowser.open(
                "https://stgit.dcs.gla.ac.uk./team-project-h/2021/cs33/cs33-main/-/blob/master/README.md",
                0,
            ),
        )
        menubar.add_cascade(label="Help", menu=helpmenu)

        # Add menubars to self.window
        self.window.config(menu=menubar)

    def setup_layout(self):
        """ Defines the layout and places widget on the TK Window """

        # Create the title for the window
        tk.Label(self.window, text="VascuSens Blockage Visualiser",
                 font=("Ubuntu", 25)).grid(row=0, columnspan=4)

        # Configure TKFrame for Side Controls
        side_controls = tk.Frame(self.window, padx=30, pady=30)
        side_controls.columnconfigure(0, weight=1)
        side_controls.columnconfigure(1, weight=2)
        side_controls.columnconfigure(1, weight=1)
        side_controls.grid(row=2, column=0, rowspan=5, sticky="n")

        # Set a heading for the side controls
        side_controls_label = ttk.Label(side_controls, text="-- VIS CONTROLS --", font="Ubuntu")
        side_controls_label.grid(row=0, column=0, sticky="n", columnspan=3)

        # Blank label here so row doesn't collapse. Used for spacing
        blank_label = tk.Label(side_controls)
        blank_label.grid(row=1, column=0)

        # Place button that allows users to upload files to generate visualisation from
        tk.Button(
            side_controls,
            text="Upload Items",
            height=2,
            width=20,
            command=self.upload_action,
            font="Ubuntu",
        ).grid(row=2, column=0, columnspan=3, sticky="we")

        # Define text variables for the file selection buttons
        self.button1_stringvar = StringVar()
        self.button2_stringvar = StringVar()
        self.button1_stringvar.set(BUTTON1_DEFAULT_TEXT)
        self.button2_stringvar.set(BUTTON2_DEFAULT_TEXT)

        # Define corresponding text variables to hold the path of the file selected
        self.button1_stringvar_path = StringVar()
        self.button2_stringvar_path = StringVar()

        # Define buttons 1 and 2 that hold the file name of whatever the user selects
        button1 = tk.Button(
            side_controls,
            textvariable=self.button1_stringvar,
            compound=tk.LEFT,
            font="Ubuntu",
            padx=8,
            pady=15,
            command=lambda: self.clear_or_replace_msgbox(1),
        )
        button1.grid(row=3, column=0, columnspan=3, sticky="we")

        button2 = tk.Button(
            side_controls,
            textvariable=self.button2_stringvar,
            compound=tk.LEFT,
            font="Ubuntu",
            padx=8,
            pady=15,
            command=lambda: self.clear_or_replace_msgbox(2),
        )
        button2.grid(row=4, column=0, columnspan=3, sticky="we")

        # Frequency Selection Options
        self.slider_value = tk.IntVar()
        self.slider_value.set(20)

        # Place the frequency slider with the callback function changedSliderValue
        freq_choice = ttk.Scale(side_controls, from_=20, to=100, orient="horizontal",
                                command=self.update_slider_value, variable=self.slider_value,)
        freq_choice.grid(row=5, column=1, sticky="we", padx=5)

        # Label for the frequency slider
        freq_label = tk.Label(side_controls, text="Frequency: ")
        freq_label.grid(row=5, column=0)

        # Show the numerical value of the slider
        self.freq_slider_val = tk.Label(side_controls, text=str(self.get_slider_value()))
        self.freq_slider_val.grid(row=5, column=2)

        # Set a label for the f_std field
        f_std_label = ttk.Label(side_controls, text="Flatten: ")
        f_std_label.grid(row=6, column=0)

        # Define variable and entrybox for the f_std field
        self.f_std_val = tk.DoubleVar()
        self.f_std_val.set(2.5)
        self.f_std = ttk.Entry(side_controls, textvariable=self.f_std_val)
        self.f_std.grid(row=6, column=1, columnspan=1)

        # Define variable and Checkbutton for the f_bool field
        self.f_bool = tk.BooleanVar()
        self.flatten_checkbutton = ttk.Checkbutton(
            side_controls,
            variable=self.f_bool,
            onvalue=True,
            offvalue=False,
        )
        self.flatten_checkbutton.grid(row=6, column=2, columnspan=1)

        # Add vertical space between the certain sections
        side_controls.grid_rowconfigure(7, minsize=15)

        # Define and place two buttons for the user to reset the visualisations
        reset_vis1 = ttk.Button(side_controls, text="Reset V. 1", command=self.reset_left)
        reset_vis1.grid(row=8, column=0, columnspan=1, sticky="w", ipadx=20)

        reset_vis2 = ttk.Button(side_controls, text="Reset V. 2", command=self.reset_right)
        reset_vis2.grid(row=8, column=1, columnspan=2, sticky="e", ipadx=20)

        # Create the generate button for the visualisation
        generate_button = tk.Button(self.window, text="Generate...", font=("Ubuntu", 20),
                                    command=self.generate, fg="green")
        generate_button.grid(row=8, column=0, ipady=25, columnspan=4, sticky="we")

        # Placeholder for the left-side visualisation object
        self.visualisation_placeholder1 = tk.Canvas(self.window, bg="#98B4D4")
        self.visualisation_placeholder1.create_text(
            150,
            10,
            fill="darkblue",
            font="Times 10 italic bold",
            text="EMPTY VISUALISATION OBJECT",
        )
        self.visualisation_placeholder1.grid(
            row=2,
            column=2,
            rowspan=5,
            sticky="nswe",
        )

        # Placeholder for the right-side visualisation object
        self.visualisation_placeholder2 = tk.Canvas(self.window, bg="#98B4D4")
        self.visualisation_placeholder2.create_text(
            150,
            10,
            fill="darkblue",
            font="Times 10 italic bold",
            text="EMPTY VISUALISATION OBJECT.",
        )
        self.visualisation_placeholder2.grid(row=2, column=3, rowspan=5, sticky="nswe", padx=(0, 10))

        # The frames that hold the visualisations
        self.left_vis_frame = tk.Frame(self.window)
        self.right_vis_frame = tk.Frame(self.window)
        self.options1 = tk.Frame(self.window)
        self.options2 = tk.Frame(self.window)

        # Define an empty baseline path until it is set
        self.baseline_path = BASELINE_DEFAULT_TEXT

    def upload_action(self, close_popup=False, keep_one=False, keep_two=False):
        """
        Handles the uploading of the visualisation files.

        Parameters
        ----------
        closePopup : boolean
            By default this is false, if this is set to true it means the call
            has come from the edit options menu, so that popup needs to close first.
        keepOne : boolean
            By default this is false, if this is set to true it means the user
            is replacing the second button item and we want to retain the first item.
        keepTwo : boolean
            By default this is false, if this is set to true it means the user is replacing
            the first button item and we want to retain the second item.
        """

        # Destroy edit menu popup, if this function is called from there
        global message_window
        if close_popup is True:
            message_window.destroy()

        # Tkinter function for selection of files
        files = filedialog.askopenfilenames()

        # Checks that only one or two files were selected
        if len(files) > 2:
            # Raise an error that only max 2 files are allowed
            messagebox.showerror("Error", "You may only select max 2 excel files.")
        else:
            # Check if the keep parameters have been specified in the function call
            # If we don't need to retain any of the button items, reset them both for the new files
            if keep_one is False and keep_two is False:
                self.button1_stringvar.set(BUTTON1_DEFAULT_TEXT)
                self.button2_stringvar.set(BUTTON2_DEFAULT_TEXT)
                self.button1_stringvar_path.set("")
                self.button2_stringvar_path.set("")
            # If we need to retain the first file item, only reset the second item
            elif keep_one is True and keep_two is False:
                self.button2_stringvar.set(BUTTON2_DEFAULT_TEXT)
                self.button2_stringvar_path.set("")
            # If we need to retain the second file item, only reset the first item
            elif keep_one is False and keep_two is True:
                self.button1_stringvar.set(BUTTON1_DEFAULT_TEXT)
                self.button1_stringvar_path.set("")

            # Check how many files the user selected from the file dialog
            if len(files) == 1:
                # Work out whether the one selected file should have a 1. or 2. prefix
                if keep_one is True:  # if this is set to true it must be a 2. prefix
                    buttontext = str("2. " + files[0].split("/")[-1])
                else:  # Else it must be a 1. prefix
                    buttontext = str("1. " + files[0].split("/")[-1])

                # Work out which button item to change for the new file
                if keep_one is False and keep_two is False:
                    self.button1_stringvar.set(buttontext)
                    self.button1_stringvar_path.set(files[0])
                elif keep_one is True:  # dont change button1_stringvar
                    self.button2_stringvar.set(buttontext)
                    self.button2_stringvar_path.set(files[0])
                elif keep_two is True:  # dont change button2_stringvar
                    self.button1_stringvar.set(buttontext)
                    self.button1_stringvar_path.set(files[0])

            # If the user selects 2 files from the upload dialog
            elif len(files) == 2:
                # Reset both buttons as there is only max 2 files allowed
                self.button1_stringvar.set(BUTTON1_DEFAULT_TEXT)
                self.button2_stringvar.set(BUTTON2_DEFAULT_TEXT)
                self.button1_stringvar_path.set("")
                self.button2_stringvar_path.set("")

                # Set the first button text for the first file
                buttontext = "1. " + files[0].split("/")[-1]
                self.button1_stringvar.set(buttontext)
                self.button1_stringvar_path.set(files[0])

                # Set the second button text for the first file
                buttontext = "2. " + files[1].split("/")[-1]
                self.button2_stringvar.set(buttontext)
                self.button2_stringvar_path.set(files[1])

    def visualisation(
        self,
        input_path,
        freq,
        forget_vis1,
        forget_vis2,
        flatten,
        flatten_std,
    ):
        """
        Handles the creation of the visualisations. Only generates one visualisation at a time.

        Parameters
        ----------
        input_path : String
            This variable specifies the file path from which to create the visualisation from.
        freq : int
            This variable specifies the integer selected from the sliding scale to use for the visualisation.
        forget_vis1 : boolean
            If this variable is true, it will delete the first visualisation
            placeholder for the generated visualisation.
        forget_vis2 : boolean
            If this variable is true, it will delete the second visualisation
            placeholder for the generated visualisation.
        flatten : boolean
            This variable is a value specified in the GUI for use in the greit_visualisation function.
        flatten_std : float
            This variable is a value specified in the GUI for use in the greit_visualisation function.
        """

        # Open input and baseline data files
        if self.baseline_path != BASELINE_DEFAULT_TEXT:
            input_data, baseline_data = open_file_at_frequency(input_path, baseline_path=self.baseline_path, freq=freq)
        else:
            input_data, baseline_data = open_file_at_frequency(input_path, freq=freq)

        # Generate the visualisation
        if flatten:
            fig = greit_visualisation(input_data, baseline_data, flatten_std)
        else:
            fig = greit_visualisation(input_data, baseline_data)

        # Check which visualisation we are replacing
        if forget_vis1:
            # Remove the visualisation placeholder from grid
            self.visualisation_placeholder1.grid_forget()
            # Place our visualisation actual holder
            self.left_vis_frame.grid(row=2, column=2, rowspan=5, sticky="we")

            # Draw our generated visualisation to grid
            visual1 = FigureCanvasTkAgg(fig, self.left_vis_frame)
            visual1.draw()
            self.options1.grid(row=7, column=2)

            # Add a save button for the visualisation
            ttk.Button(
                self.options1, text="Save Visualisation", command=lambda: self.save_vis(fig), width=45
            ).grid(row=0, column=0)

            # Place our visualisation in holder
            visual1 = visual1.get_tk_widget().grid(row=0, column=0)

        # Elif is acceptable here as this function only generates one visualisation at a time
        elif forget_vis2:
            # Remove the 2nd visualisation placeholder from grid
            self.visualisation_placeholder2.grid_forget()

            # Place our actual frame holder for the visualisation
            self.right_vis_frame.grid(row=2, column=3, rowspan=5, sticky="we", padx=10)

            # Draw our visualisation to grid
            visual2 = FigureCanvasTkAgg(fig, self.right_vis_frame)
            visual2.draw()
            self.options2.grid(row=7, column=3)

            # Add a save button for the visualisation
            ttk.Button(
                self.options2, text="Save Visualisation", command=lambda: self.save_vis(fig), width=45
            ).grid(row=0, column=0)

            # Place our visualisation in holder
            visual2 = visual2.get_tk_widget().grid(row=0, column=0)

    def save_vis(self, fig):
        """
        Handles the saving of the visualisations.

        Parameters
        ----------
        fig : matplotlib.figure.Figure
            The matplotlib figure to be saved
        """

        # Tk dialog to ask where to save the file. Allows for jpg, png and svg format but png by default
        path = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=(
                ("Portable Network Graphics", ".png"),
                ("JPEG", ".jpg"),
                ("Scalable Vector Graphics", ".svg"),
            ),
        )

        # Write to file and save
        fig.savefig(path)

    def generate(self):
        """ This function coordinates all the data ready for visualisation. Called when the user clicks the Generate
        button """

        # Perform validation on the configurations the user has set. If 0 is returned we know there is errors
        if(self.perform_validation() == 0):
            # Exit function as there is errors. No need for error message as it is handled in performValidation()
            return 0

        # Get the fbool value the user has set and tests if it is "True" to generate an actual boolean
        fbool = self.f_bool.get()

        # Get the fstd value the user has set
        fstd = self.f_std_val.get()

        # Check which section needs a visualisation generated for
        if (self.button1_stringvar.get() != BUTTON1_DEFAULT_TEXT
            and self.button2_stringvar.get() == BUTTON2_DEFAULT_TEXT):  # noqa: E129
            # If the second option is still "2. No Item Selected" then we only create a visualisation from the first
            # file
            self.visualisation(
                self.button1_stringvar_path.get(),
                self.get_slider_value(),
                True,
                False,
                fbool,
                fstd,
            )

        # If both button items have different text, then produce a visualisation for each
        elif (self.button1_stringvar.get() != BUTTON1_DEFAULT_TEXT
              and self.button2_stringvar.get() != BUTTON2_DEFAULT_TEXT):
            # Generate 2 visualisations with the relevant data from each
            self.visualisation(
                self.button1_stringvar_path.get(),
                self.get_slider_value(),
                True,
                False,
                fbool,
                fstd,
            )
            self.visualisation(
                self.button2_stringvar_path.get(),
                self.get_slider_value(),
                False,
                True,
                fbool,
                fstd,
            )

        # If the first button is "1. No Item Selected" then no need to create a visualisation from this. Only create
        # one for the 2nd if it isn't the default text
        elif (self.button1_stringvar.get() == BUTTON1_DEFAULT_TEXT
              and self.button2_stringvar.get() != BUTTON2_DEFAULT_TEXT):
            # Produce 1 visualisation from the second file with the relevant data
            self.visualisation(
                self.button2_stringvar_path.get(),
                self.get_slider_value(),
                False,
                True,
                fbool,
                fstd,
            )

    def perform_validation(self):
        """
        Performs validation on the changeable configuration for the visualisation.

        Returns
        -------
            0 : Means there is errors in the user configuration
            1 : No error in user configuration and visualisation can proceed
        """

        # If the first file has been set, check it is a file
        if self.button1_stringvar.get() != BUTTON1_DEFAULT_TEXT:
            try:
                openpyxl.load_workbook(self.button1_stringvar_path.get())
            except openpyxl.utils.exceptions.InvalidFileException:
                # Not a valid file, throw an error
                messagebox.showerror("Error", "File 1 is not an Excel File")
                return 0

        # If the second file has been set, check it is a file
        if self.button2_stringvar.get() != BUTTON2_DEFAULT_TEXT:
            try:
                openpyxl.load_workbook(self.button2_stringvar_path.get())
            except openpyxl.utils.exceptions.InvalidFileException:
                # Not a valid file, throw an error
                messagebox.showerror("Error", "File 2 is not an Excel File")
                return 0

        # Check that baseline_path is valid
        if self.baseline_path != "None Set":
            try:
                openpyxl.load_workbook(self.baseline_path)
            except openpyxl.utils.exceptions.InvalidFileException:
                # Not a valid file, throw an error
                messagebox.showerror("Error", "Baseline file is not an Excel File")
                return 0

        # Check that flatten is a float value
        try:
            fstdvaltest = self.f_std.get()
            fstdvaltest = float(fstdvaltest)
            if fstdvaltest < 0:
                raise ValueError
        except ValueError:
            # Not a float value, throw an error
            messagebox.showerror("Error", "f_std field must be a valid positive float value. ")
            return 0

        # Check that fstd is either True or False as on some systems you can type in a dropdown bo
        fbooltest = self.f_bool.get()
        if(fbooltest not in (True, False)):
            # Data in field is not from dropdown box, throw an error
            messagebox.showerror("Error", "f_bool field must be set to True or False. ")
            return 0

        # If program has got to this point there are no errors, return a 1
        return 1

    def clear_file1(self):
        """ Used to clear the set data for file1. Called from the edit menu popup when editing a file """

        # Destroy the popup window that called this function
        global message_window
        message_window.destroy()

        # If the button is already the default, there is nothing to clear
        if self.button1_stringvar.get() == BUTTON1_DEFAULT_TEXT:
            return

        # Provide a confirmation box for clearing the file
        conf_box_message = (
            "Are you sure you want to clear file: " + self.button1_stringvar.get().split(" ")[1]
        )
        answer = messagebox.askyesno("Confirmation", conf_box_message)

        # Clear the file by resetting the button text and file path
        if answer is True:
            self.button1_stringvar.set(BUTTON1_DEFAULT_TEXT)
            self.button1_stringvar_path.set("")
            message_window.destroy()
        else:
            # If user has not confirmed then exit
            return

    def clear_file2(self):
        """ Used to clear the set data for file2. Called from the edit menu popup when editing a file """

        # If the button is already the default, there is nothing to clear
        if self.button2_stringvar.get() == BUTTON2_DEFAULT_TEXT:
            return

        # Provide a confirmation box for clearing the file
        conf_box_message = (
            "Are you sure you want to clear file: " + self.button2_stringvar.get().split(" ")[1]
        )
        answer = messagebox.askyesno("Confirmation", conf_box_message)

        # Clear the file by resetting the button text and file path
        if answer is True:
            self.button2_stringvar.set(BUTTON2_DEFAULT_TEXT)
            self.button2_stringvar_path.set("")
            message_window.destroy()
        else:
            # If user has not confirmed then exit
            return

    def clear_or_replace_msgbox(self, opt):
        """
        Provide a popup for an edit menu on an existing uploaded file. Called when the user clicks on a button that
        contains the file they have selected.

        Parameters
        ----------
        opt : int
            1: indicates the first button has been pressed, so the messagebox should edit the item in the first button.
            2: indicates the second button has been pressed, so the messagebox should edit the item in the second
            button.
        """

        # Access the tk global message_window so we can create our own pop up boxes
        global message_window

        # If user has selected the first file
        if opt == 1:
            # If the default text for the first button already exists, then introduce a special pop-up box that only
            # has the option of uploading a file
            if self.button1_stringvar.get() == BUTTON1_DEFAULT_TEXT:
                # Create popup
                message_window = tk.Toplevel()

                # Add heading to popup
                tk.Label(message_window, text="Add File 1", font=("ariel 22 bold")).grid(
                    row=0, column=0, columnspan=3
                )

                # Add an upload button to messagebox that will just upload to file1
                tk.Button(
                    message_window,
                    text="Upload File",
                    command=lambda: self.upload_action(True, False, True),
                ).grid(row=1, column=1, padx=20, pady=20)

                # Add a close popup button to the window
                tk.Button(
                    message_window,
                    text="Close Window",
                    command=lambda: message_window.destroy(),
                ).grid(row=1, column=2, padx=20, pady=20)

            # If the default text doesn't exist on the button, there is an actual button to edit so provide edit
            # options such as replacing and clearing
            else:
                # Create a new pop up box
                message_window = tk.Toplevel()

                # Lay a canvas to stretch the pop up box
                popupsize = tk.Canvas(message_window, width=350, height=100)
                popupsize.grid(columnspan=3, rowspan=3)

                # Set a heading for editing of the file
                # Set popup box title to the file name
                message_window.title("Edit: " + self.button1_stringvar.get().split("1. ")[1])
                label_text = StringVar()
                label_text.set("Edit Options: " + self.button1_stringvar.get().split("1. ")[1])
                tk.Label(
                    message_window, textvariable=label_text, font=("ariel 22 bold")
                ).grid(row=0, column=0, columnspan=3)

                # Add a clear file button
                tk.Button(message_window, text="Clear File", command=self.clear_file1).grid(
                    row=1, column=0
                )

                # Add a replace file button that replaces the first file
                tk.Button(
                    message_window,
                    text="Replace File",
                    command=lambda: self.upload_action(True, False, True),
                ).grid(row=1, column=1)

                # Add a close popup button
                tk.Button(
                    message_window,
                    text="Close Window",
                    command=lambda: message_window.destroy(),
                ).grid(row=1, column=2)

        # Produce a popup window for editing of the second file
        elif opt == 2:

            # If the default text is there, there is no file to edit so only show options for uploading a file
            if self.button2_stringvar.get() == BUTTON2_DEFAULT_TEXT:
                # Create a new pop up box
                message_window = tk.Toplevel()

                # Popup box heading
                tk.Label(message_window, text="Add File 2", font=("ariel 22 bold")).grid(
                    row=0, column=0, columnspan=3
                )

                # Place an upload file button
                tk.Button(
                    message_window,
                    text="Upload File",
                    command=lambda: self.upload_action(True, True, False),
                ).grid(row=1, column=1, padx=20, pady=20)

                # Place a close window button
                tk.Button(
                    message_window,
                    text="Close Window",
                    command=lambda: message_window.destroy(),
                ).grid(row=1, column=2)

            # If the default text is not there, then there is a file to be edited so we need features such as replacing
            # and clearing
            else:
                # Spawn a new popup window
                message_window = tk.Toplevel()

                # Lay a canvas to stretch to desired size
                popupsize = tk.Canvas(message_window, width=350, height=100)
                popupsize.grid(columnspan=3, rowspan=3)

                # Set the title to the current file selected name
                message_window.title("Edit: " + self.button2_stringvar.get().split("2. ")[1])
                label_text = StringVar()
                label_text.set("Edit Options: " + self.button2_stringvar.get().split("2. ")[1])
                tk.Label(
                    message_window, textvariable=label_text, font=("ariel 22 bold")
                ).grid(row=0, column=0, columnspan=3)

                # Place a button that can clear the selected file
                tk.Button(
                    message_window, text="Clear File", command=self.clear_file2
                ).grid(row=1, column=0)

                # Place a button that replaces the current file
                tk.Button(
                    message_window,
                    text="Replace File",
                    command=lambda: self.upload_action(True, True, False),
                ).grid(row=1, column=1)

                # Place a button that closes the popup window
                tk.Button(
                    message_window,
                    text="Close Window",
                    command=lambda: message_window.destroy(),
                ).grid(row=1, column=2)

    def get_slider_value(self):
        """ This function acts as a callback required to dynamically get the tkinter slider widget value """

        # Access and return string variable dynamically and not statically
        return self.slider_value.get()

    def update_slider_value(self, freq):
        """ This function is called by tkinter when the user slides the slider for frequency selection """

        # Encapsulate in a try statement as tkinter slider prone to erroring
        try:
            # Set the text value of the slider to correspond to the scale widget selected value
            self.freq_slider_val.configure(text=str(int(float(freq))))
        except NameError:
            pass

    def reset_left(self):
        """ Reset any produced visualisations on the left side and reinstate the visualisation placeholder """

        # Forget actual visualisation holders from UI
        self.options1.grid_forget()
        self.left_vis_frame.grid_forget()

        # Reinstate visualisation placeholder to stop column collapsing
        self.visualisation_placeholder1 = tk.Canvas(self.window, bg="#98B4D4")
        self.visualisation_placeholder1.create_text(
            150,
            10,
            fill="darkblue",
            font="Times 10 italic bold",
            text="EMPTY VISUALISATION OBJECT",
        )
        self.visualisation_placeholder1.grid(
            row=2,
            column=2,
            rowspan=5,
            sticky="nswe",
        )

    def reset_right(self):
        """ Reset any produced visualisations on the right side and reinstate the visualisation placeholder """

        # Forget actual-visualisation holders from UI
        self.options2.grid_forget()
        self.right_vis_frame.grid_forget()

        # Reinstate visualisation placeholder to stop column collapsing
        self.visualisation_placeholder2 = tk.Canvas(self.window, bg="#98B4D4")
        self.visualisation_placeholder2.create_text(
            150,
            10,
            fill="darkblue",
            font="Times 10 italic bold",
            text="EMPTY VISUALISATION OBJECT.",
        )
        self.visualisation_placeholder2.grid(
            row=2, column=3, rowspan=5, sticky="nswe", padx=(0, 10)
        )

    def edit_baseline_path(self):
        """ Handle the changing of configurations for the baseline file data path by producing a pop-up box """

        # Access message_window from tk to generate a new popup
        global message_window

        # Produce a new popup window
        message_window = tk.Toplevel()

        # Set header
        tk.Label(message_window, text="Edit Baseline File Path", font=("ariel 22 bold")).grid(
            row=0, column=0, columnspan=3
        )

        # Define a string var for tk label to show current path
        self.baselinefilepath_stringvar = tk.StringVar()

        # Set to the current set path
        self.baselinefilepath_stringvar.set(self.baseline_path)

        # Show current path on the screen
        self.baselinefilepath_label = tk.Label(message_window, textvariable=self.baselinefilepath_stringvar, font=(
            "ariel 12")).grid(row=1, column=0, padx=10, pady=5)

        # Place a button that the user can press to change the path
        tk.Button(
            message_window,
            text="Change Path",
            height=2,
            width=20,
            command=self.set_baseline_path,
            font="Ubuntu",
        ).grid(row=1, column=1)

    def set_baseline_path(self):
        """ Changes the file path for the baseline file data """

        # Spawns a tk file dialog to get user to open new baseline file
        newpath = filedialog.askopenfilenames()

        # Change string var so label changes, try statement in-case the user clicks cancel on the popup window
        try:
            self.baselinefilepath_stringvar.set(newpath[0])

            # Change baseline path variable
            self.baseline_path = newpath[0]
        except IndexError:
            pass


def main():
    # Run GUI
    program = VascuSensGUI()
    program.window.mainloop()


if __name__ == "__main__":
    main()
